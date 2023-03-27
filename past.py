# get data from 도매꾹

# Initializing
import json
import requests
import openpyxl as xls
import pandas as pd

# Set up API parameters
url = 'https://domeggook.com/ssl/api/'
no = ['9426438','33960724','13863388','7544607','14422170','30905595','31793021','21279303','33607871','8364494']
params = {
    'ver': '4.4',
    'mode': 'getItemView',
    'aid': 'd761ff40c92e29464d2cc34351dae62a',
    'no': no,
    'om': 'json'
}

for num in params['no']:
    # Getting API Response
    res = requests.get(url, params=params)

    # Parsing JSON response
    data = json.loads(res.content)


    #Option askii to Korean
    data['domeggook']['selectOpt'] = eval(data['domeggook']['selectOpt'].replace('\\"','"'))

    #otherItem delete
    del data['domeggook']['desc']['contents']['otherItem']

    # Saving JSON data to a file
    with open('test.json', 'wb') as f:
        f.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))





#---------------------------------------------------------------------------------------------------
#get header and data from test.xls

# read data from excel file
dom_title_sum = pd.read_excel(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_to_godo.xlsx')

# extract header row as list
header_list = list(dom_title_sum.columns)



#1:1매칭 key:value data만 새로 new_dict만들기
def get_all_keys(d):
    result = {}
    for k, v in d.items():
        if isinstance(v, dict):
            nested_result = get_all_keys(v)
            result.update(nested_result)
        else:
            if not isinstance(k, dict) and not isinstance(v, dict):
                result[k] = v
    return result
new_dict = get_all_keys(data)



# 붙여넣을 data 생성 1row
row_list_value = []

switch = False
for head in header_list:
    for k, v in new_dict.items():
        if head.count(k) > 0:
            row_list_value.append(v)
            switch = True
            break
    if switch == False:
        row_list_value.append("")
    elif switch == True:
        switch = False



# 기존 엑셀파일에 새로운 데이터 삽입하기
wb = xls.load_workbook(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_to_godo.xlsx')
ws = wb.active

# insert values into cells
column = 1
for value in row_list_value:
    if isinstance(value, list): # value가 list인지 확인
        value = ", ".join(map(str,value)) # 리스트들 안에 있는 원소들을 있는 쉼표로 구분하여 출력
    ws.cell(row=5, column=column).value = value
    column += 1

# fixed values into cells
ws['B5'].value = int(10)
ws['AH5'].value = int(10)

# save a modified file
wb.save(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\test.xlsx')


