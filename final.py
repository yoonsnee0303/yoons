#get data from 도매꾹

# Initializing
import json
import requests
import pandas as pd
import openpyxl as xls
import time

# 도매매 상품리스트
from get_product_numbers import get_product_numbers

# # Enter your login credentials and the domeggok_url of the page to retrieve product numbers from
# username = 'dfgagu'
# password = 'df1051184!@'
# domeme_url = 'https://domemedb.domeggook.com/index/item/safeDbList.php?pageLimit=1000' 



# # Call the function to retrieve the product numbers
# product_numbers = get_product_numbers(username, password, domeme_url)


# # Setting domeggok_url
# domeggok_url = 'https://domeggook.com/ssl/api/'

# print(len(product_numbers))
# print(len(product_numbers))
# print(len(product_numbers))
# print(len(product_numbers))
# time.sleep(1)

err_cnt = 0
cnt = 0
row = 5 # 입력을 시작하는 row

# for num in product_numbers:

# # Setting Request Parameters
#     params = {
#         'ver': '4.4',
#         'mode': 'getItemView',
#         'aid': 'd761ff40c92e29464d2cc34351dae62a',
#         'no': num,
#         'om': 'json'
#     }

# # Getting API Response
#     res = requests.get(domeggok_url, params=params)

pass_point = False

# # Parsing JSON response
#     data = json.loads(res.content)

import json

# Open the JSON file
with open('data.json',encoding='utf-8') as file:
    # Load the contents of the file
    data = json.load(file)

    # print(data)
    for key in data.keys():

        if key == 'errors':
            pass_point = True

        else:
                for key in data['domeggook'].keys():
                    if key == 'selectOpt':

                        # Option askii to Korean
                        data['domeggook']['selectOpt'] = eval(data['domeggook']['selectOpt'].replace('\\"','"'))

                        #otherItem delete
                        del data['domeggook']['desc']['contents']['otherItem']
                
    if pass_point != True:
        # read data from excel file
        dom_title_sum = pd.read_excel(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_to_godo\dom_to_godo_rev1.xlsx')

        # extract header row as list
        header_list = list(dom_title_sum.columns)


        #1:1매칭 key:value data만 새로 new_dict만들기
        def get_all_keys(d):
            result = {}
            for k, v in d.items():
                if isinstance(v, dict):
                    nested_result = get_all_keys(v)
                    result.update(nested_result)
                else:
                    if not isinstance(k, dict) and not isinstance(v, dict):
                        result[k] = v
            return result
        new_dict = get_all_keys(data)



        #붙여넣을 data 생성 1row
        row_list_value = []
        switch = False
        for head in header_list:
            for k, v in new_dict.items():             
                if head.count(k) > 0:
                    if k == 'no':
                        v = data['domeggook']['basis']['no']
                        row_list_value.append(v)
                        switch = True
                        break
                    elif k == 'dome':
                        v = data['domeggook']['price']['dome']
                        row_list_value.append(v)
                        switch = True
                        break
                    else:
                        row_list_value.append(v)  
                        switch = True
                        break
            # print(row_list_value)
            if switch == False:
                row_list_value.append("")
            elif switch == True:
                switch = False
                break



        # insert values into cells
        wb = xls.load_workbook(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_to_godo\dom_to_godo_rev1.xlsx')
        ws = wb.active

        column = 1
        
        for value in row_list_value:
            if isinstance(value, list): # value가 list인지 확인
                value = ",".join(map(str,value)) # 리스트들 안에 있는 원소들을 있는 쉼표로 구분하여 출력  
            ws.cell(row=row, column=column).value = value

        
            if value == '':
                ws.cell(row=row, column=column).value = ws.cell(row=row-1, column=column).value 
            column += 1


        
        row += 1

        # save a modified file
        wb.save(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_to_godo\dom_to_godo_rev1.xlsx')
        
        # # Saving JSON data to a file
        with open('data.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)


# Done
  
    if pass_point == True:
        err_cnt+=1
    #     print('error:', err_cnt)
    # print('done')

    












