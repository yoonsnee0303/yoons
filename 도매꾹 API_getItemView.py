#get data from 도매꾹
#get data from 도매꾹
#get data from 도매꾹
#get data from 도매꾹

# Initializing
import json
import requests

# Setting URL
url = 'https://domeggook.com/ssl/api/'

# Setting Request Parameters
param = dict()
param['ver'] = '4.4'
param['mode'] = 'getItemView'
param['aid'] = 'd761ff40c92e29464d2cc34351dae62a'
param['no'] = '10944900'
param['om'] = 'json'

# Getting API Response
res = requests.get(url, params=param)

# Parsing JSON response
data = json.loads(res.content)

with open('test.json', 'wb') as f:
    f.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

#Option askii to Korean
#Option askii to Korean

data['domeggook']['selectOpt'] = eval(data['domeggook']['selectOpt'].replace('\\"','"'))

#otherItem delete
del data['domeggook']['desc']['contents']['otherItem']

# Saving JSON data to a file
with open('test.json', 'wb') as f:
    f.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))


import pandas as pd

# read data from excel file
dom_title_sum = pd.read_excel(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_title_sum.xlsx')

# extract header row as list
header_list = list(dom_title_sum.columns)



#1:1매칭 key:value data만 새로 new_dict만들기
def get_all_keys(d):
    result = {}
    for k, v in d.items():
        if isinstance(v, dict):
            nested_result = get_all_keys(v)
            result.update(nested_result)
        else:
            if not isinstance(k, dict) and not isinstance(v, dict):
                result[k] = v
    return result
new_dict = get_all_keys(data)



#붙여넣을 data 생성 1row
row_list_value = []

switch = False
for head in header_list:
    for k, v in new_dict.items():
        if head.count(k) > 0:
            row_list_value.append(v)
            switch = True
            break
    if switch == False:
        row_list_value.append("")
    elif switch == True:
        switch = False


import openpyxl as xls

# for num in range(1,len(row_list_value)+1):
#     ws.cell(row=5,column=num,value=row_list_value[num-1]) 

wb = xls.load_workbook(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\dom_title_sum.xlsx')
ws = wb.active
# get the last row number
last_row = ws.max_row + 1

# insert values into cells
column = 1

for value in row_list_value:
    if isinstance(value, list): # value가 list인지 확인
        value = ", ".join(map(str,value)) # 리스트들 안에 있는 원소들을 있는 쉼표로 구분하여 출력
    ws.cell(row=5, column=column).value = value
    column += 1

# fixed values into cells
ws['B5'].value = int(10)
ws['AH5'].value = int(10)

# save a modified file
    # wb.save(r'C:\Users\Data2\OneDrive\바탕 화면\윤\코드\test2.xlsx')
print(data)

















